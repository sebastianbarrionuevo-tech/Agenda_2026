"""
Excel Manager for Agenda 2026
Handles Excel file operations and synchronization with database
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict
import os
from database import AgendaDatabase


class ExcelManager:
    """Manages Excel file operations for the agenda"""
    
    def __init__(self, db: AgendaDatabase):
        """Initialize Excel Manager with database connection"""
        self.db = db
    
    def create_template(self, filename: str = "templates/agenda_template.xlsx"):
        """Create an Excel template for the agenda"""
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create Contacts sheet
        self._create_contacts_sheet(wb)
        
        # Create Events sheet
        self._create_events_sheet(wb)
        
        # Create Tasks sheet
        self._create_tasks_sheet(wb)
        
        # Save template
        wb.save(filename)
        print(f"Template created: {filename}")
    
    def _create_contacts_sheet(self, wb):
        """Create contacts sheet with headers"""
        ws = wb.create_sheet("Contactos")
        
        # Headers
        headers = ["ID", "Nombre", "Apellido", "Teléfono", "Email", "Dirección", "Notas"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 30
    
    def _create_events_sheet(self, wb):
        """Create events sheet with headers"""
        ws = wb.create_sheet("Eventos")
        
        # Headers
        headers = ["ID", "Título", "Descripción", "Fecha Inicio", "Fecha Fin", 
                  "Ubicación", "ID Contacto", "Recordatorio", "Completado"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
    
    def _create_tasks_sheet(self, wb):
        """Create tasks sheet with headers"""
        ws = wb.create_sheet("Tareas")
        
        # Headers
        headers = ["ID", "Título", "Descripción", "Prioridad", "Fecha Vencimiento", 
                  "Completado", "Fecha Completado"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 18
    
    def export_to_excel(self, filename: str = "data/agenda_export.xlsx"):
        """Export all data from database to Excel"""
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create and populate sheets
        self._create_contacts_sheet(wb)
        self._populate_contacts(wb['Contactos'])
        
        self._create_events_sheet(wb)
        self._populate_events(wb['Eventos'])
        
        self._create_tasks_sheet(wb)
        self._populate_tasks(wb['Tareas'])
        
        # Save file
        wb.save(filename)
        print(f"Data exported to: {filename}")
    
    def _populate_contacts(self, ws):
        """Populate contacts sheet with database data"""
        contacts = self.db.get_contacts()
        for contact in contacts:
            ws.append([
                contact['id'],
                contact['nombre'],
                contact['apellido'],
                contact['telefono'],
                contact['email'],
                contact['direccion'],
                contact['notas']
            ])
    
    def _populate_events(self, ws):
        """Populate events sheet with database data"""
        events = self.db.get_events()
        for event in events:
            ws.append([
                event['id'],
                event['titulo'],
                event['descripcion'],
                event['fecha_inicio'],
                event['fecha_fin'],
                event['ubicacion'],
                event['contacto_id'],
                event['recordatorio'],
                event['completado']
            ])
    
    def _populate_tasks(self, ws):
        """Populate tasks sheet with database data"""
        tasks = self.db.get_tasks()
        for task in tasks:
            ws.append([
                task['id'],
                task['titulo'],
                task['descripcion'],
                task['prioridad'],
                task['fecha_vencimiento'],
                task['completado'],
                task['fecha_completado']
            ])
    
    def import_from_excel(self, filename: str):
        """Import data from Excel to database"""
        if not os.path.exists(filename):
            print(f"File not found: {filename}")
            return
        
        wb = openpyxl.load_workbook(filename)
        
        # Import contacts
        if 'Contactos' in wb.sheetnames:
            self._import_contacts(wb['Contactos'])
        
        # Import events
        if 'Eventos' in wb.sheetnames:
            self._import_events(wb['Eventos'])
        
        # Import tasks
        if 'Tareas' in wb.sheetnames:
            self._import_tasks(wb['Tareas'])
        
        print(f"Data imported from: {filename}")
    
    def _import_contacts(self, ws):
        """Import contacts from Excel sheet"""
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:  # If nombre is not empty
                contact_id = row[0]
                if contact_id:
                    # Update existing contact
                    self.db.update_contact(
                        contact_id,
                        nombre=row[1] or "",
                        apellido=row[2] or "",
                        telefono=row[3] or "",
                        email=row[4] or "",
                        direccion=row[5] or "",
                        notas=row[6] or ""
                    )
                else:
                    # Add new contact
                    self.db.add_contact(
                        nombre=row[1],
                        apellido=row[2] or "",
                        telefono=row[3] or "",
                        email=row[4] or "",
                        direccion=row[5] or "",
                        notas=row[6] or ""
                    )
    
    def _import_events(self, ws):
        """Import events from Excel sheet"""
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and row[3]:  # If titulo and fecha_inicio are not empty
                event_id = row[0]
                if not event_id:
                    # Add new event
                    self.db.add_event(
                        titulo=row[1],
                        descripcion=row[2] or "",
                        fecha_inicio=str(row[3]),
                        fecha_fin=str(row[4]) if row[4] else "",
                        ubicacion=row[5] or "",
                        contacto_id=row[6] if row[6] else None,
                        recordatorio=row[7] if row[7] else 0
                    )
    
    def _import_tasks(self, ws):
        """Import tasks from Excel sheet"""
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:  # If titulo is not empty
                task_id = row[0]
                if not task_id:
                    # Add new task
                    self.db.add_task(
                        titulo=row[1],
                        descripcion=row[2] or "",
                        prioridad=row[3] or "Media",
                        fecha_vencimiento=str(row[4]) if row[4] else ""
                    )
