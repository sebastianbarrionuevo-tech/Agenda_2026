#!/usr/bin/env python3
"""
Script para crear un calendario en Excel para el año 2026
"""

import calendar
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def crear_calendario_2026():
    """
    Crea un archivo Excel con el calendario del año 2026
    """
    # Crear un nuevo libro de Excel
    wb = Workbook()
    wb.remove(wb.active)  # Eliminar la hoja por defecto
    
    # Configuración de colores y estilos
    color_encabezado = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    color_fin_semana = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    fuente_encabezado = Font(bold=True, size=12, color="FFFFFF")
    fuente_titulo = Font(bold=True, size=14)
    fuente_dia = Font(size=11)
    
    alineacion_centro = Alignment(horizontal="center", vertical="center")
    
    borde_delgado = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Nombres de los días de la semana en español
    dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    
    # Nombres de los meses en español
    nombres_meses = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    
    # Crear una hoja por cada mes
    for mes in range(1, 13):
        nombre_mes = nombres_meses[mes - 1]
        ws = wb.create_sheet(title=nombre_mes)
        
        # Título del mes
        ws.merge_cells('A1:G1')
        celda_titulo = ws['A1']
        celda_titulo.value = f'{nombre_mes} 2026'
        celda_titulo.font = fuente_titulo
        celda_titulo.alignment = alineacion_centro
        
        # Encabezados de los días de la semana
        for col, dia in enumerate(dias_semana, start=1):
            celda = ws.cell(row=2, column=col)
            celda.value = dia
            celda.font = fuente_encabezado
            celda.fill = color_encabezado
            celda.alignment = alineacion_centro
            celda.border = borde_delgado
        
        # Obtener el calendario del mes
        cal = calendar.monthcalendar(2026, mes)
        
        # Llenar los días del mes
        fila = 3
        for semana in cal:
            for col, dia in enumerate(semana, start=1):
                celda = ws.cell(row=fila, column=col)
                if dia != 0:
                    celda.value = dia
                    celda.font = fuente_dia
                    celda.alignment = alineacion_centro
                    celda.border = borde_delgado
                    
                    # Colorear fines de semana (Sábado y Domingo)
                    if col in [6, 7]:
                        celda.fill = color_fin_semana
                else:
                    celda.value = ''
                    celda.border = borde_delgado
            fila += 1
        
        # Ajustar el ancho de las columnas
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Ajustar la altura de las filas
        ws.row_dimensions[1].height = 30
        for row in range(2, fila):
            ws.row_dimensions[row].height = 25
    
    # Guardar el archivo
    nombre_archivo = 'Calendario_2026.xlsx'
    wb.save(nombre_archivo)
    print(f'✓ Calendario creado exitosamente: {nombre_archivo}')


if __name__ == '__main__':
    crear_calendario_2026()
