#!/usr/bin/env python3
"""
Ejemplo de uso del generador de calendario
Muestra cómo personalizar el calendario según necesidades específicas
"""

import calendar
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from crear_calendario import crear_calendario_2026

def agregar_eventos_personalizados():
    """
    Ejemplo de cómo agregar eventos personalizados al calendario
    """
    # Primero generamos el calendario básico
    crear_calendario_2026()
    
    # Luego abrimos el archivo para personalizarlo
    wb = load_workbook('Calendario_2026.xlsx')
    
    # Ejemplo: Marcar algunos días festivos en Enero
    ws_enero = wb['Enero']
    
    # Año Nuevo (1 de enero) - día 1 está en la fila 3, columna 4 (Jueves)
    celda_anio_nuevo = ws_enero.cell(row=3, column=4)
    celda_anio_nuevo.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    celda_anio_nuevo.font = Font(bold=True, color="FFFFFF")
    
    print("✓ Eventos personalizados agregados")
    
    # Guardar el archivo modificado
    wb.save('Calendario_2026_Personalizado.xlsx')
    print("✓ Calendario personalizado guardado como: Calendario_2026_Personalizado.xlsx")

def mostrar_informacion_calendario():
    """
    Muestra información útil sobre el calendario de 2026
    """
    print("\n=== Información del Calendario 2026 ===\n")
    
    # Información general
    print(f"Año: 2026")
    print(f"Tipo de año: {'Bisiesto' if calendar.isleap(2026) else 'Normal'}")
    print(f"Total de días: {366 if calendar.isleap(2026) else 365}")
    
    # Primer y último día del año
    primer_dia = date(2026, 1, 1)
    ultimo_dia = date(2026, 12, 31)
    dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    
    print(f"\nPrimer día del año: {dias_semana[primer_dia.weekday()]} (1 de enero)")
    print(f"Último día del año: {dias_semana[ultimo_dia.weekday()]} (31 de diciembre)")
    
    # Días por mes
    print("\n--- Días por mes ---")
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    for i, mes in enumerate(meses, 1):
        dias = calendar.monthrange(2026, i)[1]
        primer_dia_mes = date(2026, i, 1)
        print(f"{mes:12} - {dias} días (comienza en {dias_semana[primer_dia_mes.weekday()]})")

if __name__ == '__main__':
    print("=== Generador de Calendario 2026 ===\n")
    print("Seleccione una opción:")
    print("1. Generar calendario básico")
    print("2. Generar calendario con eventos personalizados")
    print("3. Mostrar información del año 2026")
    print("4. Hacer todo lo anterior")
    
    opcion = input("\nOpción (1-4): ").strip()
    
    if opcion == '1':
        crear_calendario_2026()
    elif opcion == '2':
        agregar_eventos_personalizados()
    elif opcion == '3':
        mostrar_informacion_calendario()
    elif opcion == '4':
        crear_calendario_2026()
        agregar_eventos_personalizados()
        mostrar_informacion_calendario()
    else:
        print("Opción no válida. Generando calendario básico...")
        crear_calendario_2026()
